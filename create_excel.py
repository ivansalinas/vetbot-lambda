from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear workbook
wb = Workbook()

# ============= HOJA 1: MENÚ GENERAL (25 SECCIONES) =============
ws1 = wb.active
ws1.title = "Menu General"

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
completed_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
pending_fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")
inprogress_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Headers
headers = ["Grupo", "Seccion", "Status", "Modulo Backend", "Frontend", "Horas", "Semana", "Notas"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Datos del menu
menu_data = [
    ("GESTION DEL BOT", "Flujos", "COMPLETO", "flujos", "Flujos.jsx", 2, "S1", "Endpoints 4/4 funcionales"),
    ("GESTION DEL BOT", "Catalogo", "COMPLETO", "catalogo", "Catalogo.jsx", 5, "S1", "50 productos + SIIGO en vivo"),
    ("GESTION DEL BOT", "Busqueda Inteligente", "PENDIENTE", "busqueda-rag", "Busqueda.jsx", 1, "S1", "Usa RAG Lambda (no editable)"),
    ("GESTION DEL BOT", "Productos Relacionados", "PENDIENTE", "relacionados", "Relacionados.jsx", 2, "S1", "Cross-sell config"),
    ("GESTION DEL BOT", "Respuestas Automaticas", "PENDIENTE", "respuestas", "Respuestas.jsx", 1.5, "S1", "Invocar generar_respuesta"),
    
    ("GESTION DE OPERACIONES", "Horarios y Avisos", "PENDIENTE", "horarios", "Horarios.jsx", 2, "S1", "DynamoDB tabla nueva"),
    ("GESTION DE OPERACIONES", "Credenciales", "PENDIENTE", "credenciales", "Credenciales.jsx", 2, "S2", "Secrets Manager"),
    ("GESTION DE OPERACIONES", "Personalidad", "PENDIENTE", "personalidad", "Personalidad.jsx", 1, "S2", "Secrets Manager"),
    ("GESTION DE OPERACIONES", "Logs y Monitoreo", "PENDIENTE", "logs", "Logs.jsx", 2, "S2", "CloudWatch + busquedas del bot"),
    
    ("CONFIGURACION", "Usuarios del Panel", "PENDIENTE", "usuarios", "Usuarios.jsx", 3, "S1", "JWT + DynamoDB"),
    ("CONFIGURACION", "Dashboard Ejecutivo", "PENDIENTE", "dashboard", "Dashboard.jsx", 2, "S1", "KPIs basicos"),
    ("CONFIGURACION", "Metricas del Bot", "PENDIENTE", "metricas", "Metricas.jsx", 2, "S2", "CloudWatch queries"),
    
    ("OPERACIONES FASE 2", "Pedidos Activos", "PENDIENTE", "pedidos", "Pedidos.jsx", 2, "S3", "DynamoDB pedidos"),
    ("OPERACIONES FASE 2", "Gestion Domiciliarios", "PENDIENTE", "domiciliarios", "Domiciliarios.jsx", 2, "S3", "Asignacion entregas"),
    ("OPERACIONES FASE 2", "Logistica y Entregas", "PENDIENTE", "logistica", "Logistica.jsx", 2, "S3", "Track entregas"),
    
    ("CRM", "Clientes y Perfiles", "PENDIENTE", "clientes", "Clientes.jsx", 2, "S3", "DynamoDB clientes"),
    ("CRM", "Historial de Compras", "PENDIENTE", "historial", "Historial.jsx", 2, "S3", "Compras + preferencias"),
    ("CRM", "Segmentacion", "PENDIENTE", "segmentacion", "Segmentacion.jsx", 2, "S4", "Grupos de clientes"),
    
    ("MARKETING", "Campanas de WhatsApp", "PENDIENTE", "campanas", "Campanas.jsx", 2, "S4", "Broadcast via Meta"),
    ("MARKETING", "Promociones y Descuentos", "PENDIENTE", "promociones", "Promociones.jsx", 2, "S4", "Cupones"),
    
    ("COMUNICACION", "Chat en Tiempo Real", "PENDIENTE", "chat", "Chat.jsx", 2, "S4", "WebSocket"),
    ("COMUNICACION", "Notificaciones", "PENDIENTE", "notificaciones", "Notificaciones.jsx", 1, "S4", "Alerts"),
    
    ("ANALISIS", "Respaldo y Restauracion", "PENDIENTE", "backup", "Backup.jsx", 2, "S4", "S3 + DynamoDB"),
    ("ANALISIS", "Reporte de Ventas", "PENDIENTE", "reportes-ventas", "Reportes.jsx", 2, "S4", "SQL queries"),
    ("ANALISIS", "Reporte de Clientes", "PENDIENTE", "reportes-clientes", "Reportes.jsx", 2, "S4", "Analytics"),
]

# Insertar datos
for row_idx, data in enumerate(menu_data, 2):
    grupo, seccion, status, backend, frontend, horas, semana, notas = data
    
    ws1.cell(row=row_idx, column=1, value=grupo)
    ws1.cell(row=row_idx, column=2, value=seccion)
    ws1.cell(row=row_idx, column=3, value=status)
    ws1.cell(row=row_idx, column=4, value=backend)
    ws1.cell(row=row_idx, column=5, value=frontend)
    ws1.cell(row=row_idx, column=6, value=horas)
    ws1.cell(row=row_idx, column=7, value=semana)
    ws1.cell(row=row_idx, column=8, value=notas)
    
    # Formato por status
    for col in range(1, 9):
        cell = ws1.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        if status == "COMPLETO":
            cell.fill = completed_fill
        elif status == "PENDIENTE":
            cell.fill = pending_fill
        elif status == "EN PROGRESO":
            cell.fill = inprogress_fill

# Ajustar anchos
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 8
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 30

# ============= HOJA 2: RESUMEN EJECUTIVO =============
ws2 = wb.create_sheet("Resumen Ejecutivo")

ws2['A1'] = "VETBOT PANEL ADMIN - CONTROL Y SEGUIMIENTO"
ws2['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws2.merge_cells('A1:D1')

# Status General
ws2['A3'] = "STATUS GENERAL"
ws2['A3'].font = Font(bold=True, size=12)

ws2['A4'] = "Modulos Completados:"
ws2['B4'] = 2
ws2['A5'] = "Modulos Pendientes:"
ws2['B5'] = 23
ws2['A6'] = "Paginas Frontend Completadas:"
ws2['B6'] = 0
ws2['A7'] = "Paginas Frontend Pendientes:"
ws2['B7'] = 5 + 20

# Horas
ws2['A9'] = "HORAS INVERTIDAS"
ws2['A9'].font = Font(bold=True, size=12)

ws2['A10'] = "Ejecutadas:"
ws2['B10'] = 10
ws2['A11'] = "Estimadas Fase 1 (Semana 1):"
ws2['B11'] = 14
ws2['A12'] = "Estimadas Fase 2 (Semanas 2-4):"
ws2['B12'] = 39
ws2['A13'] = "TOTAL:"
ws2['B13'] = 63
ws2['B13'].font = Font(bold=True)

# Proximo Paso
ws2['A15'] = "PROXIMO PASO CRITICO"
ws2['A15'].font = Font(bold=True, size=12, color="FF0000")

ws2['A16'] = "Martes 3 SEP 9 AM: Iniciar Frontend Panel Admin (Vite + React)"
ws2['A17'] = "Objetivo Viernes 6 SEP: 5 paginas + conectadas a API"
ws2['A18'] = "Deadline: Maria revisa MVP antes de Fase 2"

for row in range(16, 19):
    ws2[f'A{row}'].alignment = Alignment(wrap_text=True)

ws2.column_dimensions['A'].width = 50
ws2.column_dimensions['B'].width = 15

# Guardar
wb.save(r'C:\Users\ivanm\AWS\vetbot\VetBot_Control_Seguimiento.xlsx')
print("OK")

