"""
VetBot — Carga inicial del catálogo a S3 Vectors + DynamoDB
Lee el Excel del catálogo, genera embeddings con Bedrock Titan
y los indexa en S3 Vectors para búsqueda semántica (RAG).

USO:
    # Con datos mock (para pruebas sin Excel del cliente):
    python cargar_catalogo.py --profile vetbot-pamascotas --mock

    # Con Excel real del cliente:
    python cargar_catalogo.py --profile vetbot-pamascotas --excel catalogo.xlsx

    # Solo verificar sin cargar:
    python cargar_catalogo.py --profile vetbot-pamascotas --mock --dry-run
"""

import boto3, json, argparse, sys, time
from pathlib import Path
from botocore.exceptions import ClientError

REGION       = "us-east-1"
ACCOUNT_ID   = "330631894163"
VECTOR_BUCKET= "vetbot-pamascotas-vectors"
FOTOS_BUCKET = f"vetbot-pamascotas-fotos-{ACCOUNT_ID}"
TABLE_CAT    = "vetbot-catalogo"   # tabla DynamoDB con los datos del catálogo
EMBED_MODEL  = "amazon.titan-embed-text-v2:0"
EMBED_DIM    = 1024                # dimensiones del vector Titan V2

ok   = lambda m: print(f"✅  {m}")
info = lambda m: print(f"ℹ️   {m}")
err  = lambda m: print(f"❌  {m}")
step = lambda n,t: print(f"\n{'─'*50}\n🔧 Paso {n}: {t}\n{'─'*50}")


# ─── DATOS MOCK — 20 productos Pa'Mascotas ─────────────
MOCK_PRODUCTOS = [
    {"codigo_siigo":"RC-MAXI-15",    "nombre":"Royal Canin Maxi Adult 15kg",
     "descripcion_larga":"Alimento completo para perros adultos de razas grandes (26-44kg). Articulaciones y digestión. Bolsa 15kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Royal Canin","peso_presentacion":"15kg","nombre_foto":"RC-MAXI-15.jpg",
     "palabras_clave":"raza grande,adulto,mantenimiento","precio_mock":189000,"stock_mock":45},

    {"codigo_siigo":"RC-MAXI-LIGHT", "nombre":"Royal Canin Maxi Light 15kg",
     "descripcion_larga":"Control de peso para perros grandes con sobrepeso. Alto en fibra, bajo en calorías. Bolsa 15kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Royal Canin","peso_presentacion":"15kg","nombre_foto":"RC-MAXI-LIGHT.jpg",
     "palabras_clave":"light,sobrepeso,dieta,raza grande","precio_mock":195000,"stock_mock":32},

    {"codigo_siigo":"RC-MINI-2",     "nombre":"Royal Canin Mini Adult 2kg",
     "descripcion_larga":"Alimento para perros adultos de razas pequeñas (hasta 10kg). Croqueta pequeña adaptada. Bolsa 2kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Royal Canin","peso_presentacion":"2kg","nombre_foto":"RC-MINI-2.jpg",
     "palabras_clave":"raza pequeña,adulto,miniatura","precio_mock":38000,"stock_mock":60},

    {"codigo_siigo":"RC-PUPPY-4",    "nombre":"Royal Canin Puppy 4kg",
     "descripcion_larga":"Alimento para cachorros de todas las razas hasta los 12 meses. Fortalece el sistema inmune. Bolsa 4kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Cachorro",
     "marca":"Royal Canin","peso_presentacion":"4kg","nombre_foto":"RC-PUPPY-4.jpg",
     "palabras_clave":"cachorro,puppy,todas las razas","precio_mock":72000,"stock_mock":48},

    {"codigo_siigo":"HILLS-AD-15",   "nombre":"Hills Science Diet Adult 15kg",
     "descripcion_larga":"Fórmula veterinaria para manejo de peso en perros adultos. DHA y antioxidantes. Bolsa 15kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Hills","peso_presentacion":"15kg","nombre_foto":"HILLS-AD-15.jpg",
     "palabras_clave":"hills,veterinario,peso,adulto","precio_mock":215000,"stock_mock":20},

    {"codigo_siigo":"ACANA-AC-11",   "nombre":"Acana Classics Prairie Poultry 11.4kg",
     "descripcion_larga":"Alimento premium sin cereales. 50% ingredientes de origen animal. Pollo y huevo de granja. Bolsa 11.4kg.",
     "categoria":"Alimento perro","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Acana","peso_presentacion":"11.4kg","nombre_foto":"ACANA-AC-11.jpg",
     "palabras_clave":"premium,sin cereales,grain free,natural","precio_mock":285000,"stock_mock":15},

    {"codigo_siigo":"RC-CAT-4",      "nombre":"Royal Canin Feline Health 4kg",
     "descripcion_larga":"Alimento completo para gatos adultos en interiores. Controla la formación de bolas de pelo. Bolsa 4kg.",
     "categoria":"Alimento gato","especie":"Gato","etapa_vida":"Adulto",
     "marca":"Royal Canin","peso_presentacion":"4kg","nombre_foto":"RC-CAT-4.jpg",
     "palabras_clave":"gato adulto,interior,bolas de pelo","precio_mock":68000,"stock_mock":40},

    {"codigo_siigo":"RC-KITTEN-2",   "nombre":"Royal Canin Kitten 2kg",
     "descripcion_larga":"Alimento para gatitos desde el destete hasta los 12 meses. Con DHA para el desarrollo cerebral. Bolsa 2kg.",
     "categoria":"Alimento gato","especie":"Gato","etapa_vida":"Cachorro",
     "marca":"Royal Canin","peso_presentacion":"2kg","nombre_foto":"RC-KITTEN-2.jpg",
     "palabras_clave":"gatito,kitten,cachorro gato","precio_mock":42000,"stock_mock":35},

    {"codigo_siigo":"AGILITY-CAT-3", "nombre":"Agility Gold Gatos Adultos 3kg",
     "descripcion_larga":"Alimento balanceado para gatos adultos de todas las razas. Omega 3 y 6 para pelaje brillante. Bolsa 3kg.",
     "categoria":"Alimento gato","especie":"Gato","etapa_vida":"Adulto",
     "marca":"Agility","peso_presentacion":"3kg","nombre_foto":"AGILITY-CAT-3.jpg",
     "palabras_clave":"economico,gato adulto,pelaje","precio_mock":32000,"stock_mock":55},

    {"codigo_siigo":"DENTASTIX-L",   "nombre":"DentaStix Large Pedigree x28",
     "descripcion_larga":"Snack dental para perros grandes (+25kg). Reduce el sarro hasta un 80% con uso diario. Sabor pollo. Caja 28u.",
     "categoria":"Snacks","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Pedigree","peso_presentacion":"28 unidades","nombre_foto":"DENTASTIX-L.jpg",
     "palabras_clave":"dental,sarro,dientes,snack,premio grande","precio_mock":28000,"stock_mock":80},

    {"codigo_siigo":"DENTASTIX-M",   "nombre":"DentaStix Medium Pedigree x28",
     "descripcion_larga":"Snack dental para perros medianos (10-25kg). Limpieza dental diaria. Sabor original. Caja 28 unidades.",
     "categoria":"Snacks","especie":"Perro","etapa_vida":"Adulto",
     "marca":"Pedigree","peso_presentacion":"28 unidades","nombre_foto":"DENTASTIX-M.jpg",
     "palabras_clave":"dental,sarro,dientes,mediano","precio_mock":24000,"stock_mock":75},

    {"codigo_siigo":"WHISKAS-SNACK", "nombre":"Whiskas Temptations Salmón 85g",
     "descripcion_larga":"Snack irresistible para gatos adultos. Exterior crujiente, interior suave. Sabor salmón. Bolsa 85g.",
     "categoria":"Snacks","especie":"Gato","etapa_vida":"Adulto",
     "marca":"Whiskas","peso_presentacion":"85g","nombre_foto":"WHISKAS-SNACK.jpg",
     "palabras_clave":"snack gato,premio,salmon,temptations","precio_mock":12000,"stock_mock":100},

    {"codigo_siigo":"FRONT-L",       "nombre":"Frontline Spot-On Perros L (20-40kg)",
     "descripcion_larga":"Pipeta antiparasitaria para perros grandes. Elimina pulgas, garrapatas y piojos. Efecto 4 semanas. 1 pipeta.",
     "categoria":"Medicamentos","especie":"Perro","etapa_vida":"Todas",
     "marca":"Frontline","peso_presentacion":"1 pipeta","nombre_foto":"FRONT-L.jpg",
     "palabras_clave":"antipulgas,garrapatas,antiparasitario,grande","precio_mock":38000,"stock_mock":50},

    {"codigo_siigo":"FRONT-M",       "nombre":"Frontline Spot-On Perros M (10-20kg)",
     "descripcion_larga":"Pipeta antiparasitaria para perros medianos. Pulgas y garrapatas. Protección hasta 4 semanas. 1 pipeta.",
     "categoria":"Medicamentos","especie":"Perro","etapa_vida":"Todas",
     "marca":"Frontline","peso_presentacion":"1 pipeta","nombre_foto":"FRONT-M.jpg",
     "palabras_clave":"antipulgas,garrapatas,antiparasitario,mediano","precio_mock":34000,"stock_mock":45},

    {"codigo_siigo":"FRONT-CAT",     "nombre":"Frontline Spot-On Gatos",
     "descripcion_larga":"Pipeta antiparasitaria para gatos. Elimina pulgas y garrapatas. Uso mensual. 1 pipeta.",
     "categoria":"Medicamentos","especie":"Gato","etapa_vida":"Todas",
     "marca":"Frontline","peso_presentacion":"1 pipeta","nombre_foto":"FRONT-CAT.jpg",
     "palabras_clave":"antipulgas,garrapatas,antiparasitario gato","precio_mock":32000,"stock_mock":40},

    {"codigo_siigo":"SHAMP-DOG-250", "nombre":"Shampoo Virbac Perros 250ml",
     "descripcion_larga":"Shampoo dermatológico para perros de pelo normal. pH neutro. Suaviza y da brillo. Frasco 250ml.",
     "categoria":"Higiene","especie":"Perro","etapa_vida":"Todas",
     "marca":"Virbac","peso_presentacion":"250ml","nombre_foto":"SHAMP-DOG-250.jpg",
     "palabras_clave":"shampoo,baño,pelo,higiene,dermatologico","precio_mock":28000,"stock_mock":30},

    {"codigo_siigo":"CATSAN-10",     "nombre":"Arena Catsan Ultra 10 litros",
     "descripcion_larga":"Arena sanitaria aglutinante para gatos. Absorción máxima, controla el olor 7 días. Sin polvo. 10 litros.",
     "categoria":"Arena","especie":"Gato","etapa_vida":"Todas",
     "marca":"Catsan","peso_presentacion":"10 litros","nombre_foto":"CATSAN-10.jpg",
     "palabras_clave":"arena,sanitaria,aglutinante,olor,gato","precio_mock":45000,"stock_mock":25},

    {"codigo_siigo":"CORREA-M",      "nombre":"Correa Ajustable Nylon Mediana",
     "descripcion_larga":"Correa resistente de nylon para perros medianos. Largo 1.5m, ancho 2cm. Colores variados.",
     "categoria":"Accesorios","especie":"Perro","etapa_vida":"Todas",
     "marca":"Genérica","peso_presentacion":"1.5m","nombre_foto":"CORREA-M.jpg",
     "palabras_clave":"correa,paseo,nylon,mediana","precio_mock":22000,"stock_mock":20},

    {"codigo_siigo":"CAMA-DOG-M",    "nombre":"Cama Acolchada Perros Mediana",
     "descripcion_larga":"Cama suave y acolchada para perros medianos. Funda lavable. Talla M (60x45cm). Varios colores.",
     "categoria":"Accesorios","especie":"Perro","etapa_vida":"Todas",
     "marca":"Genérica","peso_presentacion":"60x45cm","nombre_foto":"CAMA-DOG-M.jpg",
     "palabras_clave":"cama,descanso,acolchada,mediana","precio_mock":85000,"stock_mock":10},

    {"codigo_siigo":"JUGUETE-KONG",  "nombre":"Kong Classic Relleable Talla M",
     "descripcion_larga":"Juguete de caucho natural para perros medianos. Rellenable con premios. Estimulación mental. Talla M.",
     "categoria":"Accesorios","especie":"Perro","etapa_vida":"Todas",
     "marca":"Kong","peso_presentacion":"Talla M","nombre_foto":"JUGUETE-KONG.jpg",
     "palabras_clave":"juguete,kong,entretenimiento,rellenable,caucho","precio_mock":55000,"stock_mock":18},
]


def texto_para_embedding(p: dict) -> str:
    """Construye el texto que se convierte en embedding para búsqueda semántica."""
    partes = [
        p["nombre"],
        p["descripcion_larga"],
        f"Categoría: {p['categoria']}",
        f"Especie: {p['especie']}",
        f"Etapa: {p.get('etapa_vida','Todas')}",
        f"Marca: {p.get('marca','')}",
        p.get("palabras_clave","").replace(",", " "),
    ]
    return " | ".join(filter(None, partes))


def generar_embedding(bedrock_client, texto: str) -> list[float]:
    """Llama a Bedrock Titan Embeddings V2 y devuelve el vector."""
    resp = bedrock_client.invoke_model(
        modelId=EMBED_MODEL,
        body=json.dumps({
            "inputText": texto,
            "dimensions": EMBED_DIM,
            "normalize": True,
        }),
        contentType="application/json",
        accept="application/json",
    )
    body = json.loads(resp["body"].read())
    return body["embedding"]


def crear_vector_bucket(s3_client) -> None:
    """Crea el bucket vectorial de S3 Vectors si no existe."""
    try:
        # S3 Vectors usa la API de vector-buckets
        s3_client.create_bucket(
            Bucket=VECTOR_BUCKET,
            CreateBucketConfiguration={"LocationConstraint": REGION}
            if REGION != "us-east-1" else {},
        )
        ok(f"Bucket vectorial creado: {VECTOR_BUCKET}")
    except Exception as e:
        if "BucketAlreadyOwnedByYou" in str(e) or "BucketAlreadyExists" in str(e):
            info(f"Bucket vectorial ya existe: {VECTOR_BUCKET}")
        else:
            # S3 Vectors puede tener su propia API — intentar con endpoint específico
            info(f"Nota: {e} — verificar en consola que el bucket vectorial existe")


def crear_tabla_catalogo(ddb_client) -> None:
    """Crea la tabla DynamoDB del catálogo si no existe."""
    try:
        ddb_client.create_table(
            TableName=TABLE_CAT,
            AttributeDefinitions=[
                {"AttributeName": "codigo_siigo", "AttributeType": "S"}
            ],
            KeySchema=[{"AttributeName": "codigo_siigo", "KeyType": "HASH"}],
            BillingMode="PAY_PER_REQUEST",
            Tags=[
                {"Key": "proyecto", "Value": "vetbot"},
                {"Key": "cliente",  "Value": "pamascotas"},
            ]
        )
        ok(f"Tabla DynamoDB creada: {TABLE_CAT}")
        info("Esperando que la tabla esté activa...")
        waiter = ddb_client.get_waiter("table_exists")
        waiter.wait(TableName=TABLE_CAT)
        ok("Tabla activa")
    except ClientError as e:
        if "ResourceInUseException" in str(e):
            info(f"Tabla ya existe: {TABLE_CAT}")
        else:
            raise


def cargar_producto(producto: dict, bedrock, ddb_resource, s3_client,
                    dry_run: bool = False) -> bool:
    """
    Para cada producto:
    1. Genera el embedding con Bedrock Titan
    2. Guarda el vector en S3 (como JSON — hasta que S3 Vectors esté disponible)
    3. Guarda los datos completos en DynamoDB
    """
    codigo = producto["codigo_siigo"]
    texto  = texto_para_embedding(producto)

    if dry_run:
        info(f"[DRY RUN] {codigo}: '{texto[:60]}...'")
        return True

    try:
        # 1. Generar embedding
        vector = generar_embedding(bedrock, texto)

        # 2. Guardar vector en S3 (JSON con el vector + metadatos)
        vector_obj = {
            "codigo_siigo": codigo,
            "embedding":    vector,
            "texto_fuente": texto,
            "dimension":    EMBED_DIM,
        }
        s3_client.put_object(
            Bucket=FOTOS_BUCKET,  # Usando bucket de fotos como storage temporal
            Key=f"embeddings/{codigo}.json",
            Body=json.dumps(vector_obj),
            ContentType="application/json",
        )

        # 3. Guardar datos en DynamoDB
        tabla = ddb_resource.Table(TABLE_CAT)
        item = {**producto, "texto_embedding": texto}
        # Convertir números a Decimal para DynamoDB
        from decimal import Decimal
        for k, v in item.items():
            if isinstance(v, float):
                item[k] = Decimal(str(v))
            elif isinstance(v, int):
                item[k] = Decimal(str(v))
        tabla.put_item(Item=item)

        ok(f"{codigo} — embedding generado y guardado")
        return True

    except Exception as e:
        err(f"{codigo}: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description="Carga catálogo VetBot a S3 Vectors + DynamoDB")
    parser.add_argument("--profile", default="vetbot-pamascotas")
    parser.add_argument("--mock",    action="store_true", help="Usar datos mock internos")
    parser.add_argument("--excel",   type=str, help="Ruta al Excel del catálogo real")
    parser.add_argument("--dry-run", action="store_true", help="Solo mostrar, no cargar")
    args = parser.parse_args()

    if not args.mock and not args.excel:
        err("Debes especificar --mock o --excel <ruta>")
        sys.exit(1)

    print(f"""
{'='*55}
VetBot — Carga Catálogo RAG
Profile: {args.profile} | Mock: {args.mock} | DryRun: {args.dry_run}
{'='*55}
""")

    session  = boto3.Session(profile_name=args.profile, region_name=REGION)
    cuenta   = session.client("sts").get_caller_identity()["Account"]
    if cuenta != ACCOUNT_ID:
        err(f"Cuenta incorrecta: {cuenta}"); sys.exit(1)
    ok(f"Cuenta: {cuenta}")

    bedrock  = session.client("bedrock-runtime", region_name=REGION)
    ddb      = session.client("dynamodb", region_name=REGION)
    ddb_res  = session.resource("dynamodb", region_name=REGION)
    s3       = session.client("s3", region_name=REGION)

    # Cargar productos
    if args.mock:
        productos = MOCK_PRODUCTOS
        info(f"Usando {len(productos)} productos mock")
    else:
        import openpyxl
        wb   = openpyxl.load_workbook(args.excel)
        ws   = wb.active
        hdrs = [ws.cell(row=3, column=i).value for i in range(1, 13)]
        productos = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0]:  # código_siigo no vacío
                p = {hdrs[i]: row[i] for i in range(len(hdrs)) if hdrs[i]}
                productos.append(p)
        info(f"Cargados {len(productos)} productos del Excel: {args.excel}")

    # Crear infraestructura si no existe
    step(1, "Verificar/crear tabla DynamoDB del catálogo")
    if not args.dry_run:
        crear_tabla_catalogo(ddb)

    step(2, f"Generar embeddings y cargar {len(productos)} productos")
    ok_count = 0
    for i, p in enumerate(productos, 1):
        print(f"  [{i:02d}/{len(productos)}] ", end="", flush=True)
        if cargar_producto(p, bedrock, ddb_res, s3, args.dry_run):
            ok_count += 1
        if not args.dry_run and i % 5 == 0:
            time.sleep(0.5)  # Evitar throttling de Bedrock

    print(f"""
{'='*55}
✅  Carga completada
   Productos procesados: {ok_count}/{len(productos)}
   
Próximo paso — construir el motor de búsqueda RAG:
   python motor_rag.py --profile vetbot-pamascotas --test "comida para perro con sobrepeso"
{'='*55}
""")


if __name__ == "__main__":
    main()
